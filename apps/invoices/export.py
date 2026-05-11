import io
from datetime import date
from decimal import Decimal
from typing import Iterable

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Invoice

STATUS_LABELS = dict(Invoice.STATUS_CHOICES)

COLUMNS = [
    ('Numer faktury', 30),
    ('Nr referencyjny KSeF', 38),
    ('Sprzedawca', 40),
    ('NIP sprzedawcy', 16),
    ('Data wystawienia', 18),
    ('Termin płatności', 18),
    ('Netto', 16),
    ('VAT', 14),
    ('Brutto', 16),
    ('Waluta', 8),
    ('MPP', 8),
    ('Status', 24),
    ('Konto bankowe', 30),
    ('Notatki', 40),
]

HEADER_FILL = PatternFill(start_color='1A56DB', end_color='1A56DB', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
ODD_FILL = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
)
OVERDUE_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')


def _apply_header(ws, row: int = 1):
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 30


def _invoice_to_row(inv: Invoice) -> list:
    return [
        inv.invoice_number,
        inv.ksef_reference_number,
        inv.seller_name,
        inv.seller_nip,
        inv.issue_date,
        inv.payment_due_date,
        float(inv.amount_net),
        float(inv.amount_vat),
        float(inv.amount_gross),
        inv.currency,
        'TAK' if inv.is_split_payment else 'NIE',
        STATUS_LABELS.get(inv.status, inv.status),
        inv.bank_account_number,
        inv.notes,
    ]


def generate_excel(invoices: Iterable[Invoice], title: str = 'Faktury kosztowe') -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.freeze_panes = 'A2'

    _apply_header(ws)

    today = date.today()

    for row_idx, inv in enumerate(invoices, start=2):
        row_data = _invoice_to_row(inv)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')

            if col_idx in (7, 8, 9):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            if col_idx in (5, 6) and value:
                cell.number_format = 'YYYY-MM-DD'

        # Podświetl przeterminowane
        if (inv.payment_due_date and inv.payment_due_date < today
                and inv.status not in (Invoice.STATUS_PAID, Invoice.STATUS_SENT_FOR_PAYMENT)):
            fill = OVERDUE_FILL
        elif row_idx % 2 == 0:
            fill = ODD_FILL
        else:
            fill = None

        if fill:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Suma brutto
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=8, value='SUMA BRUTTO:').font = Font(bold=True)
    ws.cell(row=last_row, column=9, value=f'=SUM(I2:I{ws.max_row - 1})').number_format = '#,##0.00'
    ws.cell(row=last_row, column=9).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
